# Imports
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from sklearn.preprocessing import PolynomialFeatures
from sklearn.pipeline import make_pipeline
from sklearn.linear_model import LinearRegression
from sklearn.preprocessing import StandardScaler
from sklearn.multioutput import MultiOutputRegressor
import seaborn as sb
from sklearn.metrics import mean_squared_error, r2_score, mean_absolute_error

# For debugging
current_method = "Polynomial Regression"
print(f"Starting exection for {current_method} method.")

# Load data
train_path = 'train_data.xlsx'
train_data = pd.read_excel(train_path, sheet_name = 'Log Transformed')
test_path = 'test_data.xlsx'
test_data =  pd.read_excel(test_path, sheet_name = None)
print("Data has been loaded.")

# Define features and target
features = [
    'passenger_count', 'trip_distance', 'pickup_location_id',
    'dropoff_location_id', 'hour', 'day', 'month'
    ]
targets = [ 'fare_amount', 'tip_amount', 'tolls_amount', 'duration',  'log_duration']
print("Features and targets have been defined.")

# Define data split
dataE = test_data[ 'Log Transformed']
x_train, y_train = train_data[features], train_data[targets]
x, y_base, y_total  = dataE[features], dataE['fare_amount'], dataE['total_amount']
y_tip, y_tolls, y_dur,  y_log_dur = dataE['tip_amount'], dataE['tolls_amount'], dataE['duration'], dataE['log_duration']
print("Data has been assigned to training and validation variables.")

# Train
model = MultiOutputRegressor(make_pipeline(StandardScaler(), PolynomialFeatures(degree = 3), LinearRegression()))
model.fit(x_train, y_train)
print(f"Training complete for {current_method}.")

# Predict
y_pred = model.predict(x)
print(f"Predictions were made for {current_method}.")

# Extract targets
predictions = {targets[i]: y_pred[:, i] for i in range(len(targets))}
print("All targets have been extracted and defined.")

# Create df
feature_df = dataE[features].copy()
target_df = pd.DataFrame(predictions)
pred_df = pd.concat([feature_df, target_df], axis = 1)
print("Prediction DataFrame has been created.")

# Calculate total amount
pred_df['total_amount'] = (pred_df['fare_amount'] + pred_df['tip_amount'] + pred_df['tolls_amount'] + 0.30) # Add surcharge back for accuracy
print("All targets have been defined.")

# Save to excel
book = load_workbook(test_path)
if "PR Predicted" in book.sheetnames:
    del book["PR Predicted"]
    book.save(test_path)
with pd.ExcelWriter(test_path, engine = "openpyxl", mode = "a") as writer:
    pred_df.to_excel(writer, sheet_name = "PR Predicted", index = False)
print(f"All predictions using {current_method} have been saved to excel.")

# Performace Metrics
def evaluate_performance(y_true, y_pred, label=""):
    mse = mean_squared_error(y_true, y_pred)
    rmse = np.sqrt(mse)
    r2 = r2_score(y_true, y_pred)
    mae = mean_absolute_error(y_true, y_pred)
    print(f"\n{label} MSE: {mse:.3f}")
    print(f"{label} RMSE: {rmse:.3f}")
    print(f"{label} R2: {r2:.3f}")
    print(f"{label} MAE: {mae:.3f}")
    return mse, r2

# Base fare and total amount
print(f"\nPerformance metrics have been calculated for {current_method}:")
mse_base, r2_base = evaluate_performance(y_base, pred_df['fare_amount'], label="Fare")
mse_tips, r2_tips= evaluate_performance(y_tip, pred_df['tip_amount'], label="Tips")
mse_tolls, r2_tolls = evaluate_performance(y_tolls, pred_df['tolls_amount'], label="Tolls")
mse_dur, r2_dur = evaluate_performance(y_dur, pred_df['duration'], label="Duration")
mse_dur, r2_dur = evaluate_performance(y_log_dur, pred_df['log_duration'], label="Log Transformed Duration")
mse_total, r2_total = evaluate_performance(y_total, pred_df['total_amount'], label="Total")

# Base fare plot
plt.scatter(y_base, pred_df['fare_amount'], color = 'blue', label = 'Predictions')
plt.plot([min(y_base), max(y_base)], [min(y_base), max(y_base)], color = 'red', linestyle = '--', label = 'Perfect Prediction')
plt.xlabel('Actual Fare (USD)')
plt.ylabel('Predicted Fare (USD)')
plt.title('Fare: Actual vs Predicted (Test Data)')
plt.legend()
plt.show()

# Total amount plot
plt.scatter(y_total, pred_df['total_amount'], color = 'blue', label = 'Predictions')
plt.plot([min(y_total), max(y_total)], [min(y_total), max(y_total)], color = 'red', linestyle = '--', label = 'Perfect Prediction')
plt.xlabel('Actual Total (USD)')
plt.ylabel('Predicted Total (USD)')
plt.title('Total Paid: Actual vs Predicted (Test Data)')
plt.legend()
plt.show()
print(f"Plots created. Completed execution for {current_method}.")
